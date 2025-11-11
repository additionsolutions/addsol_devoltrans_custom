import frappe
from frappe import _
import pandas as pd
from io import BytesIO
from frappe.utils.file_manager import save_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


# --------------------------------------------------------------------------- #
#  DOWNLOAD BOM TEMPLATE
# --------------------------------------------------------------------------- #
@frappe.whitelist()
def download_bom_template():
    """
    Final template:
      - BOM sheet: rows 1-2 blank, row 3 operations, row 4 blank, row 5 headers, row 6 sample
      - SubAssy1: same layout
      - Validation Data sheet: UOM, Item Group, Operation (visible)
      - Dropdowns (data validation) for UOM and Item Group in BOM/SubAssy1 (rows 6..1000)
    """

    # ----------------- Safe master fetch -----------------
    try:
        uoms = frappe.get_all("UOM", pluck="name") or []
    except Exception:
        uoms = []

    try:
        item_groups = frappe.get_all("Item Group", pluck="name") or []
    except Exception:
        item_groups = []

    # For operations, try common fields
    try:
        operations = frappe.get_all("Operation", pluck="name") or []
    except Exception:
        try:
            # fallback in case field is different: pluck 'operation_name'
            operations = frappe.get_all("Operation", pluck="operation_name") or []
        except Exception:
            operations = []

    # If no operations found, use sensible placeholders so template shows something
    if not operations:
        operations = ["Cutting", "Machining", "Welding", "Assembly", "Inspection"]

    # ----------------- Create Workbook -----------------
    wb = Workbook()

    # ---------- Helper: styling ----------
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    center = Alignment(horizontal="center", vertical="center")

    # ----------------- Instructions sheet -----------------
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    ws_instr["A1"] = "BOM Upload Template"
    ws_instr["A1"].font = Font(bold=True, size=13)
    ws_instr["A2"] = "Instructions for BOM Upload Template"
    ws_instr["A4"] = "This Excel file defines Bill of Materials (BOM) in a structured, recursive format. Each sheet corresponds to one BOM, and sub-assemblies are linked using the “Sub BOM Sheet” column."
    ws_instr["A5"] = ""
    ws_instr["A6"] = "General Rules:"
    ws_instr["A7"] = "1. Do not rename or delete the first sheet ('Instructions')."
    ws_instr["A8"] = "2. Every BOM must have a unique sheet name (e.g. 'BOM', 'SubAssy1', 'SubAssy2')."
    ws_instr["A9"] = "3. The main BOM must always be in the sheet named 'BOM'."
    ws_instr["A10"] = "4. Operations must be listed in row 3, one per column, in the order they appear in production."
    ws_instr["A11"] = "5. The BOM data must start from row 5, where column headers are defined exactly as below."
    ws_instr["A12"] = "6. You can define nested BOMs by specifying the sub-BOM sheet name in the 'Sub BOM Sheet' column."
    ws_instr["A13"] = "7. If an Item does not exist in the system, it will be automatically created using the provided details."
    ws_instr["A14"] = "8. If any new Operation is not found, upload will fail — operations must exist beforehand."
    ws_instr["A16"] = "Column Definitions (from Row 5):"
    ws_instr["A16"].font = Font(bold=True)
    ws_instr["A17"] = "Item Code — Unique code for item (e.g. MTR-001)."
    ws_instr["A18"] = "Item Name — Readable item name (e.g. Motor Assembly)."
    ws_instr["A19"] = "Description — Short technical description."
    ws_instr["A20"] = "Make — Manufacturer or brand (e.g. Siemens, SKF)."
    ws_instr["A21"] = "Part no. — Vendor part number or internal reference."
    ws_instr["A22"] = "Qty — Quantity required for this BOM."
    ws_instr["A23"] = "UOM — Unit of Measure (e.g. Nos, Set, Kg)."
    ws_instr["A24"] = "Remarks — Additional manufacturing or purchase notes." 
    ws_instr["A25"] = "Specification Link — Optional URL to a drawing/spec sheet."
    ws_instr["A26"] = "Sub BOM Sheet — If the item is a sub-assembly, put the sheet name of its BOM (must exist as another sheet)."
    ws_instr["A28"] = "Example Workflow:"
    ws_instr["A28"].font = Font(bold=True)
    ws_instr["A29"] = "- Open the 'BOM' sheet and fill your main item and its components."
    ws_instr["A30"] = "- If any component itself has a BOM, note its sheet name in 'Sub BOM Sheet'."
    ws_instr["A31"] = "- Create a new sheet with that name (e.g. 'SubAssy1') and define its operations + components in the same format."
    ws_instr["A32"] = "- Save and upload the file from ERPNext’s 'Upload BOM Excel' function."
    ws_instr["A33"] = ""
    ws_instr["A34"] = "Important: Do not alter header capitalization, add extra columns, or change row positions (operations = row 3, headers = row 5)."
    ws_instr["A35"] = ""
    
    # ----------------- Create Sheets ------------------------
    ws = wb.create_sheet("BOM")
    sub = wb.create_sheet("SubAssy1")
    # Create Validation Sheet and protect it
    ws_val = wb.create_sheet("Validation Data")
    ws_val.protection.sheet = True
    ws_val.protection.enable()
    
    # ----------------- Validation Data sheet -----------------
    ws_val.append(["UOM", "Item Group", "Operation"])
    max_len = max(len(uoms), len(item_groups), len(operations), 1)
    for i in range(max_len):
        ws_val.append([
            uoms[i] if i < len(uoms) else "",
            item_groups[i] if i < len(item_groups) else "",
            operations[i] if i < len(operations) else ""
        ])
    # Bold header
    for cell in ws_val[1]:
        cell.font = header_font

    # ----------------- BOM sheet -----------------
    # Rows 1-2 blank
    ws.append(["BOM Sheet"])
    ws.append([])

    # Bold header row (row 1)
    for cell in ws[1]:
        cell.font = header_font

    # operations populated from masters (or placeholders)
    # Row 3: add "Operations" label + operation names
    ops_row = ["Operations: "] + operations
    ws.append(ops_row)

    # Make the first cell ("Operations") bold
    ws["A3"].font = Font(bold=True)

    # Row 4 blank
    ws.append([])

    # Row 5 headers exactly as expected by upload script
    headers = [
        "Item Code", "Item Name", "Description", "Make", "Part no.",
        "Qty", "UOM", "Item Group", "Remarks", "Specification Link", "Sub BOM Sheet"
    ]
    ws.append(headers)  # row 5

    # Row 6 sample data
    ws.append([
        "FG-001", "Motor Assembly", "Main motor assembly", "Siemens", "S-MTR-100",
        1, "Nos", "Finished Goods", "Final product", "http://example.com/specs/motor.pdf", ""
    ])

    # Row 7 sample data for sub BOM
    ws.append([
        "FG-SA-004", "Motor Sub Assembly", "Sub motor assembly", "", "AB-98-DF",
        1, "Nos", "Sub Assemblies", "Semi finished product", "http://example.com/specs/motor.pdf", "SubAssy1"
    ])

    # Bold header row (row 5)
    for cell in ws[5]:
        cell.font = header_font

    # Autosize columns
    for col in ws.columns:
        try:
            max_len_col = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(10, max_len_col + 2)
        except Exception:
            pass

    # ----------------- SubAssy1 sheet (nested BOM example) -----------------
    # rows 1-2 blank
    sub.append([])
    sub.append([])

    # row 3: operations for subassembly (reuse same operations)
    # add "Operations" label + operation names
    sub_ops_row = ["Operations: "] + operations
    sub.append(sub_ops_row)
    # Make the first cell ("Operations") bold
    sub["A3"].font = Font(bold=True)

    # row 4 blank
    sub.append([])

    # row 5 headers (same as BOM)
    sub.append(headers)

    # sample row 6
    sub.append([
        "SUB-001", "Stator Sub-Assembly", "Stator unit", "Siemens", "S-ST-010",
        1, "Nos", "Sub Assembly", "Refer SubAssy1", "http://example.com/specs/stator.pdf", ""
    ])

    # Bold headers row 5
    for cell in sub[5]:
        cell.font = header_font

    # Autosize columns on sub
    for col in sub.columns:
        try:
            max_len_col = max(len(str(c.value or "")) for c in col)
            sub.column_dimensions[get_column_letter(col[0].column)].width = max(10, max_len_col + 2)
        except Exception:
            pass

    # ----------------- Data Validations (UOM, Item Group, Operation) -----------------
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    # Determine list lengths (ensure non-empty to avoid invalid ranges)
    uom_count = max(1, len(uoms))
    ig_count = max(1, len(item_groups))
    op_count = max(1, len(operations))

    # Define explicit ranges on Validation Data sheet
    uom_range = f"'Validation Data'!$A$2:$A${uom_count + 1}"
    ig_range = f"'Validation Data'!$B$2:$B${ig_count + 1}"
    op_range = f"'Validation Data'!$C$2:$C${op_count + 1}"

    # Remove previous definitions if script re-runs
    for name in ["UOM_List", "ItemGroup_List", "Operation_List"]:
        if name in wb.defined_names:
            del wb.defined_names[name]

    # Register workbook-level named ranges (Excel will recognize these as dropdown sources)
    wb.defined_names.add(DefinedName("UOM_List", attr_text=uom_range))
    wb.defined_names.add(DefinedName("ItemGroup_List", attr_text=ig_range))
    wb.defined_names.add(DefinedName("Operation_List", attr_text=op_range))

    # --- Helper to attach validations ---
    def apply_validations(ws):
        # Each DataValidation must be a fresh instance per sheet
        uom_dv = DataValidation(
            type="list",
            formula1="=UOM_List",
            allow_blank=False,
            promptTitle="Select UOM",
            prompt="Choose a valid Unit of Measure (UOM)."
        )
        ig_dv = DataValidation(
            type="list",
            formula1="=ItemGroup_List",
            allow_blank=False,
            promptTitle="Select Item Group",
            prompt="Choose an Item Group from the Validation Data sheet."
        )
        op_dv = DataValidation(
            type="list",
            formula1="=Operation_List",
            allow_blank=False,
            promptTitle="Select Operation",
            prompt="Choose an Operation from the Validation Data sheet."
        )

        # Attach validations to sheet
        ws.add_data_validation(uom_dv)
        ws.add_data_validation(ig_dv)
        ws.add_data_validation(op_dv)

        # Apply to column ranges
        uom_dv.add("G6:G1000")    # UOM dropdown
        ig_dv.add("H6:H1000")     # Item Group dropdown
        op_dv.add("A3:ZZ3")       # Operations dropdown row

    # Apply to both BOM and SubAssy1
    apply_validations(ws)
    apply_validations(sub)

    # ----------------- Finalize and return file -----------------
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    file_bytes = bio.getvalue()

    filename = "BOM_Template.xlsx"

    # Optional: save to File for auditing but don't fail if it errors
    try:
        from frappe.utils.file_manager import save_file
        save_file(
            fname=filename,
            content=file_bytes,
            dt=None,
            dn=None,
            folder="Home/Attachments",
            is_private=1,
            decode=False
        )
        frappe.db.commit()
    except Exception as e:
        frappe.log_error(f"Failed to save BOM template file: {e}")

    # Return binary download
    frappe.response["filename"] = filename
    frappe.response["filecontent"] = file_bytes
    frappe.response["type"] = "download"
    frappe.response["content_type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
